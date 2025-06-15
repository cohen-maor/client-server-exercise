from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os

EXCEL_FILE = "live_resource_summary.xlsx"

def flatten_dict(d, parent_key='', sep='_'):
    """Flatten nested dictionary for Excel row representation."""
    items = []
    for k, v in d.items():
        new_key = f"{parent_key}{sep}{k}" if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        else:
            items.append((new_key, v))
    return dict(items)

def save_or_update_socket_summary(socket_id, summary_dict):
    flat_data = flatten_dict(summary_dict)
    headers = ["socket_id"] + list(flat_data.keys())

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)  # First row: column names
    else:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active

        # Add any new keys that didn’t exist before
        existing_headers = [cell.value for cell in ws[1]]
        new_keys = [k for k in flat_data.keys() if k not in existing_headers]
        if new_keys:
            for key in new_keys:
                existing_headers.append(key)
                ws.cell(row=1, column=len(existing_headers), value=key)
            headers = existing_headers  # Update working headers

    # Find row for socket_id
    socket_column = 1  # socket_id is always in column 1
    row_num = None
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=socket_column).value == socket_id:
            row_num = row
            break

    if row_num is None:
        row_num = ws.max_row + 1
        ws.cell(row=row_num, column=socket_column, value=socket_id)

    # Write data
    for col_idx, header in enumerate(headers[1:], start=2):  # Start at col 2
        ws.cell(row=row_num, column=col_idx, value=flat_data.get(header))

    wb.save(EXCEL_FILE)